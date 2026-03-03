#!/usr/bin/env python3
import json
from pathlib import Path

try:
    import openpyxl
except Exception as exc:  # pragma: no cover
    raise SystemExit(f"openpyxl is required to extract CoFID: {exc}")

ROOT = Path(__file__).resolve().parents[1]
DATA_DIR = ROOT / "data" / "cofid"
OUT_FILE = DATA_DIR / "cofid_compiled.json"

FILES = [
    ("cofid_2021.xlsx", "cofid2021"),
    ("cofid_oldfoods.xlsx", "cofidold"),
]

SHEET_CODE_FIELDS = [
    (
        "1.3 Proximates",
        {
            "Water (g)": "water_g",
            "Protein (g)": "protein_g",
            "Fat (g)": "fat_g",
            "Carbohydrate (g)": "carbs_g",
            "Energy (kcal) (kcal)": "energy_kcal",
            "Starch (g)": "starch_g",
            "Total sugars (g)": "sugars_g",
        },
    ),
    (
        "1.4 Inorganics",
        {
            "Sodium (mg)": "sodium_mg",
            "Potassium (mg)": "potassium_mg",
            "Calcium (mg)": "calcium_mg",
            "Magnesium (mg)": "magnesium_mg",
            "Phosphorus (mg)": "phosphorus_mg",
            "Iron (mg)": "iron_mg",
            "Copper (mg)": "copper_mg",
            "Zinc (mg)": "zinc_mg",
            "Manganese (mg)": "manganese_mg",
            "Selenium (µg)": "selenium_ug",
        },
    ),
    (
        "1.5 Vitamins",
        {
            "Retinol Equivalent (µg)": "vitamin_a_ug",
            "Vitamin D (µg)": "vitamin_d_ug",
            "Vitamin E (mg)": "vitamin_e_mg",
            "Vitamin K1 (µg)": "vitamin_k_ug",
            "Thiamin (mg)": "thiamin_b1_mg",
            "Riboflavin (mg)": "riboflavin_b2_mg",
            "Niacin equivalent (mg)": "vitamin_b3_mg",
            "Vitamin B6 (mg)": "vitamin_b6_mg",
            "Vitamin B12 (µg)": "vitamin_b12_ug",
            "Folate (µg)": "folate_ug",
        },
    ),
]

FAT_SHEETS = {
    "1.8 (SFA per 100gFood)": {"total_code": "sat_fat_g"},
    "1.10 (MUFA per 100gFood)": {"total_code": "monounsaturated_g"},
    "1.12 (PUFA per 100gFood)": {
        "total_code": "polyunsaturated_g",
        "omega3_code": "omega3_g",
        "omega6_code": "omega6_g",
    },
}

ALLOWED_CODES = {
    "energy_kcal",
    "water_g",
    "protein_g",
    "fat_g",
    "carbs_g",
    "starch_g",
    "sugars_g",
    "sat_fat_g",
    "monounsaturated_g",
    "polyunsaturated_g",
    "omega3_g",
    "omega6_g",
    "sodium_mg",
    "potassium_mg",
    "calcium_mg",
    "magnesium_mg",
    "phosphorus_mg",
    "iron_mg",
    "copper_mg",
    "zinc_mg",
    "manganese_mg",
    "selenium_ug",
    "vitamin_a_ug",
    "vitamin_d_ug",
    "vitamin_e_mg",
    "vitamin_k_ug",
    "thiamin_b1_mg",
    "riboflavin_b2_mg",
    "vitamin_b3_mg",
    "vitamin_b6_mg",
    "vitamin_b12_ug",
    "folate_ug",
}


def to_number(value):
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    raw = str(value).strip()
    if not raw:
        return None
    if raw.lower() == "n":
        return None
    if raw.lower() == "tr":
        return 0.0
    cleaned = raw.replace(",", "").replace("<", "").replace(">", "")
    try:
        return float(cleaned)
    except ValueError:
        return None


def ensure_food(foods, tag, code, name):
    key = f"{tag}:{code}"
    food = foods.get(key)
    if food is None:
        food = {
            "barcode": key,
            "name": " ".join(str(name or "").split()),
            "source": "cofid",
            "locale": "en-gb",
            "nutrients": {},
        }
        foods[key] = food
    elif name and not food.get("name"):
        food["name"] = " ".join(str(name or "").split())
    return food


def parse_workbook(filepath, tag):
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    foods = {}

    for sheet_name, code_map in SHEET_CODE_FIELDS:
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if len(rows) < 4:
            continue
        headers = [str(x or "").strip() for x in rows[0]]
        col_map = {idx: code_map[h] for idx, h in enumerate(headers) if h in code_map}

        for row in rows[3:]:
            code = str(row[0]).strip() if row and row[0] else ""
            if not code:
                continue
            food = ensure_food(foods, tag, code, row[1] if len(row) > 1 else "")
            for idx, nutrient_code in col_map.items():
                if nutrient_code not in ALLOWED_CODES:
                    continue
                n = to_number(row[idx] if idx < len(row) else None)
                if n is None:
                    continue
                food["nutrients"][nutrient_code] = n

    for sheet_name, fat_spec in FAT_SHEETS.items():
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if len(rows) < 4:
            continue
        headers = [str(x or "").strip().lower() for x in rows[0]]

        for row in rows[3:]:
            code = str(row[0]).strip() if row and row[0] else ""
            if not code:
                continue
            food = ensure_food(foods, tag, code, row[1] if len(row) > 1 else "")
            total = 0.0
            omega3 = 0.0
            omega6 = 0.0
            saw = False
            for idx in range(7, len(row)):
                n = to_number(row[idx])
                if n is None:
                    continue
                saw = True
                total += n
                h = headers[idx] if idx < len(headers) else ""
                if fat_spec.get("omega3_code") and "n-3" in h:
                    omega3 += n
                if fat_spec.get("omega6_code") and "n-6" in h:
                    omega6 += n
            if saw:
                food["nutrients"][fat_spec["total_code"]] = total
            if fat_spec.get("omega3_code") and omega3 > 0:
                food["nutrients"][fat_spec["omega3_code"]] = omega3
            if fat_spec.get("omega6_code") and omega6 > 0:
                food["nutrients"][fat_spec["omega6_code"]] = omega6

    return [f for f in foods.values() if f.get("name") and f.get("nutrients")]


def main():
    all_foods = []
    for filename, tag in FILES:
        fp = DATA_DIR / filename
        if not fp.exists():
            print(f"skip {filename} (not found)")
            continue
        print(f"extracting {filename}...")
        parsed = parse_workbook(fp, tag)
        print(f"parsed {len(parsed)} foods from {filename}")
        all_foods.extend(parsed)

    OUT_FILE.parent.mkdir(parents=True, exist_ok=True)
    OUT_FILE.write_text(json.dumps({"foods": all_foods}, ensure_ascii=False), encoding="utf-8")
    print(f"wrote {OUT_FILE} ({len(all_foods)} foods)")


if __name__ == "__main__":
    main()
